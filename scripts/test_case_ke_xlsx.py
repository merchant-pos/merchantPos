"""Mengubah TEST-CASE-KAATAGO.md jadi berkas Excel yang bisa dipakai menguji.

Test case bukan bahan bacaan seperti FSD dan TSD — ia daftar kerja.
Penguji perlu menyaring per prioritas, mengurutkan per modul, dan
menuliskan hasilnya di sebelah tiap baris. Dokumen Word tidak bisa
melakukan satu pun dari itu.

Karena itu keluarannya menambahkan tiga kolom yang tidak ada di
markdown-nya: Hasil, Catatan, dan Diuji Oleh. Kolom Hasil punya daftar
pilihan (Lulus / Gagal / Terblokir / Dilewati) supaya isinya seragam —
"OK", "ok", dan "Oke" di tiga baris berbeda membuat rekapnya harus
dibaca satu per satu.

Pakai:
    /usr/bin/python3 scripts/test_case_ke_xlsx.py
"""

import os
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SUMBER = os.path.join(REPO, "docs", "TEST-CASE-KAATAGO.md")
TUJUAN = os.path.join(REPO, "docs", "TEST-CASE-KAATAGO.xlsx")

# Folder Drive tempat tangkapan layar temuan dikumpulkan.
FOLDER_CAPTURE = (
    "https://drive.google.com/drive/folders/"
    "1MonfUjlqV0aABrLDleDZhugd3pULLIOf?usp=sharing"
)

UNGU = "4F46E5"
UNGU_MUDA = "EEEDFC"
ABU = "F1F5F9"
PUTIH = "FFFFFF"
MERAH = "B91C1C"
KUNING = "B45309"
HIJAU = "047857"

WARNA_P = {"P1": MERAH, "P2": KUNING, "P3": HIJAU}

tepi = Side(style="thin", color="D8DBE8")
KOTAK = Border(left=tepi, right=tepi, top=tepi, bottom=tepi)


def bersih(teks):
    """Membuang penanda markdown supaya selnya terbaca sebagai kalimat."""
    teks = re.sub(r"\*\*(.+?)\*\*", r"\1", teks)
    teks = re.sub(r"`(.+?)`", r"\1", teks)
    return teks.strip()


def baca(path):
    """Membaca tiap tabel kasus uji berikut bab tempatnya berada.

    Yang diambil hanya tabel berkolom lima yang barisnya dimulai TC-.
    Tabel lain di dokumen itu — prioritas, prasyarat, keterlacakan —
    memang bukan kasus uji, dan menyeretnya ikut masuk akan membuat
    penghitungan di kolom rekap salah tanpa terlihat salah.
    """
    bab = None
    hasil = []
    for baris in open(path).read().split("\n"):
        if baris.startswith("## "):
            bab = re.sub(r"^##\s+\d+\.\s*", "", baris).strip()
        if not baris.startswith("| TC-"):
            continue
        sel = [bersih(s) for s in baris.strip().strip("|").split(" | ")]
        if len(sel) != 5:
            continue
        tc, p, skenario, harapan, rujukan = sel
        hasil.append((bab, tc, p, skenario, harapan, rujukan))
    return hasil


def e2e(path):
    """Alur ujung-ke-ujung ditulis sebagai daftar bernomor, bukan tabel.

    Langkahnya digabung jadi satu sel bermultibaris — dipecah jadi satu
    baris per langkah akan memberi kesan tiap langkah punya hasilnya
    sendiri, padahal yang dinilai adalah alurnya sebagai satu kesatuan.
    """
    isi = open(path).read()
    blok = re.findall(r"### (E2E-\d+) — (.+?) \((P\d)\)\n(.*?)(?=\n### |\n---)",
                      isi, re.S)
    hasil = []
    for kode, judul, p, badan in blok:
        langkah = [bersih(b) for b in re.findall(r"^\d+\. (.+)$", badan, re.M)]
        if not langkah:
            langkah = [bersih(b) for b in badan.strip().split("\n") if b.strip()
                       and not b.startswith("**")]
        rujuk = re.search(r"\*\*Rujukan:\*\* (.+)", badan)
        hasil.append((
            "Uji Ujung-ke-Ujung", kode, p, bersih(judul),
            "\n".join(f"{i+1}. {l}" for i, l in enumerate(langkah)),
            bersih(rujuk.group(1)) if rujuk else "",
        ))
    return hasil


def judul_kolom(ws, kolom, lebar):
    for i, (nama, w) in enumerate(zip(kolom, lebar), start=1):
        sel = ws.cell(row=1, column=i, value=nama)
        sel.font = Font(bold=True, color=PUTIH, size=11)
        sel.fill = PatternFill("solid", fgColor=UNGU)
        sel.alignment = Alignment(vertical="center", horizontal="left",
                                  wrap_text=True)
        sel.border = KOTAK
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 26
    ws.freeze_panes = "A2"


def lembar_kasus(wb, kasus):
    ws = wb.create_sheet("Kasus Uji")
    kolom = ["Modul", "ID", "P", "Skenario / Langkah",
             "Hasil yang Diharapkan", "Rujukan",
             "Hasil", "Catatan", "Diuji Oleh"]
    judul_kolom(ws, kolom, [22, 12, 6, 52, 52, 20, 13, 34, 16])

    for baris in kasus:
        ws.append(list(baris) + ["", "", ""])

    for r in range(2, ws.max_row + 1):
        for c in range(1, len(kolom) + 1):
            sel = ws.cell(row=r, column=c)
            sel.border = KOTAK
            sel.alignment = Alignment(vertical="top", wrap_text=True)
        ws.cell(row=r, column=3).font = Font(
            bold=True, color=WARNA_P.get(ws.cell(row=r, column=3).value, "000000"))
        ws.cell(row=r, column=3).alignment = Alignment(
            vertical="top", horizontal="center")
        ws.cell(row=r, column=2).font = Font(bold=True)
        if r % 2 == 0:
            for c in range(1, len(kolom) + 1):
                ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=ABU)

    # Daftar pilihan supaya isinya seragam. Tanpa ini, rekapnya harus
    # dibaca satu per satu untuk tahu "ok" dan "OK" itu hal yang sama.
    dv = DataValidation(
        type="list",
        formula1='"Lulus,Gagal,Terblokir,Dilewati"',
        allow_blank=True,
        showDropDown=False,
    )
    ws.add_data_validation(dv)
    dv.add(f"G2:G{ws.max_row}")

    ws.auto_filter.ref = f"A1:I{ws.max_row}"
    return ws


def lembar_rekap(wb, ws_kasus, kasus):
    ws = wb.create_sheet("Rekap", 0)
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14

    ws["A1"] = "MerchantPOS — Rekap Pengujian"
    ws["A1"].font = Font(bold=True, size=15, color=UNGU)
    ws["A2"] = "Angkanya menghitung sendiri dari lembar Kasus Uji."
    ws["A2"].font = Font(size=10, italic=True, color="6E728C")

    n = ws_kasus.max_row
    kepala = ["Kelompok", "Jumlah", "Lulus", "Gagal", "Terblokir",
              "Dilewati", "Belum diuji"]
    for i, teks in enumerate(kepala, start=1):
        sel = ws.cell(row=4, column=i, value=teks)
        sel.font = Font(bold=True, color=PUTIH)
        sel.fill = PatternFill("solid", fgColor=UNGU)
        sel.border = KOTAK
        sel.alignment = Alignment(horizontal="center")

    def rumus(baris, kriteria_kolom, kriteria, hasil=None):
        if hasil is None:
            return f'=COUNTIF(\'Kasus Uji\'!{kriteria_kolom}2:{kriteria_kolom}{n},"{kriteria}")'
        return (f'=COUNTIFS(\'Kasus Uji\'!{kriteria_kolom}2:{kriteria_kolom}{n},"{kriteria}",'
                f"'Kasus Uji'!G2:G{n},\"{hasil}\")")

    r = 5
    for p in ["P1", "P2", "P3"]:
        ws.cell(row=r, column=1, value=f"Prioritas {p}")
        ws.cell(row=r, column=2, value=rumus(r, "C", p))
        for i, h in enumerate(["Lulus", "Gagal", "Terblokir", "Dilewati"],
                              start=3):
            ws.cell(row=r, column=i, value=rumus(r, "C", p, h))
        ws.cell(row=r, column=7,
                value=f'=B{r}-SUM(C{r}:F{r})')
        r += 1

    ws.cell(row=r, column=1, value="TOTAL").font = Font(bold=True)
    for c in range(2, 8):
        L = get_column_letter(c)
        ws.cell(row=r, column=c, value=f"=SUM({L}5:{L}{r-1})").font = Font(bold=True)
    total_baris = r
    r += 2

    ws.cell(row=r, column=1, value="Per modul").font = Font(bold=True)
    r += 1
    for i, teks in enumerate(kepala, start=1):
        sel = ws.cell(row=r, column=i, value=teks)
        sel.font = Font(bold=True, color=PUTIH)
        sel.fill = PatternFill("solid", fgColor=UNGU)
        sel.border = KOTAK
        sel.alignment = Alignment(horizontal="center")
    r += 1

    for modul in dict.fromkeys(k[0] for k in kasus):
        ws.cell(row=r, column=1, value=modul)
        ws.cell(row=r, column=2, value=rumus(r, "A", modul))
        for i, h in enumerate(["Lulus", "Gagal", "Terblokir", "Dilewati"],
                              start=3):
            ws.cell(row=r, column=i, value=rumus(r, "A", modul, h))
        ws.cell(row=r, column=7, value=f"=B{r}-SUM(C{r}:F{r})")
        r += 1

    for baris in list(range(5, total_baris + 1)) + list(range(total_baris + 4, r)):
        for c in range(1, 8):
            sel = ws.cell(row=baris, column=c)
            sel.border = KOTAK
            if c > 1:
                sel.alignment = Alignment(horizontal="center")

    ws.cell(row=r + 1, column=1,
            value="Rilis ditahan selama masih ada P1 yang Gagal.").font = Font(
        italic=True, color=MERAH)
    return ws, r + 3


def rekap_temuan(ws, mulai, awal, akhir):
    """Blok ringkasan temuan, dihitung dari lembar Defect."""
    r = mulai
    ws.cell(row=r, column=1, value="Temuan").font = Font(bold=True, size=12,
                                                        color=UNGU)
    r += 1
    kepala = ["Severity", "Jumlah", "Baru", "Dikonfirmasi",
              "Sedang diperbaiki", "Menunggu retest", "Ditutup"]
    for i, teks in enumerate(kepala, start=1):
        sel = ws.cell(row=r, column=i, value=teks)
        sel.font = Font(bold=True, color=PUTIH)
        sel.fill = PatternFill("solid", fgColor=UNGU)
        sel.border = KOTAK
        sel.alignment = Alignment(horizontal="center", wrap_text=True)
    r += 1

    rentang_sev = f"Defect!M{awal}:M{akhir}"
    rentang_st = f"Defect!P{awal}:P{akhir}"
    for sev in ["Blocker", "Major", "Minor", "Trivial"]:
        ws.cell(row=r, column=1, value=sev)
        ws.cell(row=r, column=2, value=f'=COUNTIF({rentang_sev},"{sev}")')
        for i, st in enumerate(["Baru", "Dikonfirmasi", "Sedang diperbaiki",
                                "Menunggu retest", "Ditutup"], start=3):
            ws.cell(row=r, column=i,
                    value=f'=COUNTIFS({rentang_sev},"{sev}",{rentang_st},"{st}")')
        r += 1

    ws.cell(row=r, column=1, value="TOTAL").font = Font(bold=True)
    for c in range(2, 8):
        L = get_column_letter(c)
        ws.cell(row=r, column=c,
                value=f"=SUM({L}{r-4}:{L}{r-1})").font = Font(bold=True)

    for baris in range(mulai + 1, r + 1):
        for c in range(1, 8):
            sel = ws.cell(row=baris, column=c)
            sel.border = KOTAK
            if c > 1:
                sel.alignment = Alignment(horizontal="center")

    ws.cell(row=r + 2, column=1,
            value="Blocker yang belum Ditutup menahan rilis.").font = Font(
        italic=True, color=MERAH)


def lembar_prasyarat(wb, path):
    ws = wb.create_sheet("Prasyarat")
    judul_kolom(ws, ["#", "Prasyarat", "Kenapa", "Siap?"], [8, 58, 62, 10])
    isi = open(path).read()
    for baris in isi.split("\n"):
        if not baris.startswith("| L-"):
            continue
        sel = [bersih(s) for s in baris.strip().strip("|").split(" | ")]
        if len(sel) == 3:
            ws.append(sel + [""])
    for r in range(2, ws.max_row + 1):
        for c in range(1, 5):
            k = ws.cell(row=r, column=c)
            k.border = KOTAK
            k.alignment = Alignment(vertical="top", wrap_text=True)
        ws.cell(row=r, column=1).font = Font(bold=True)
    dv = DataValidation(type="list", formula1='"Ya,Belum"', allow_blank=True,
                        showDropDown=False)
    ws.add_data_validation(dv)
    dv.add(f"D2:D{ws.max_row}")
    return ws


# Kolom lembar Defect. Urutannya mengikuti urutan orang mengisinya saat
# menemukan sesuatu — identitas dulu, lalu apa yang terjadi, baru
# penanganannya. Formulir yang meminta "Ditugaskan ke" sebelum "Apa yang
# terjadi" membuat orang berhenti mengisi di tengah.
KOLOM_DEFECT = [
    ("ID", 10, None),
    ("Tanggal", 13, None),
    ("Pelapor", 16, None),
    ("Versi APK", 13, None),
    ("Perangkat / OS", 20, None),
    ("Modul", 20, None),
    ("Peran / Akun", 16, None),
    ("TC Terkait", 14, None),
    ("Ringkasan", 44, None),
    ("Langkah Reproduksi", 52, None),
    ("Hasil Aktual", 40, None),
    ("Hasil Diharapkan", 40, None),
    ("Severity", 14, "Blocker,Major,Minor,Trivial"),
    ("Prioritas", 12, "P1,P2,P3"),
    ("Frekuensi", 14, "Selalu,Kadang,Sekali saja"),
    ("Status", 16, "Baru,Dikonfirmasi,Sedang diperbaiki,Menunggu retest,"
                   "Ditutup,Ditolak,Duplikat"),
    ("Ditugaskan ke", 16, None),
    ("Versi Perbaikan", 15, None),
    ("Tanggal Ditutup", 15, None),
    ("Link Capture", 40, None),
    ("Buka", 10, None),
    ("Catatan", 34, None),
]

BARIS_KOSONG = 60


def lembar_defect(wb):
    """Daftar temuan, berikut tombol pembuka folder capture di Drive.

    Gambarnya tidak ditempel ke dalam berkas ini. Satu tangkapan layar
    HP berukuran 300-800 KB; tiga puluh temuan sudah cukup membuat
    berkas Excel-nya terlalu berat untuk dikirim lewat chat mana pun —
    dan berkas yang tidak bisa dikirim tidak dipakai siapa-siapa.
    """
    ws = wb.create_sheet("Defect")

    ws["A1"] = "Daftar Temuan"
    ws["A1"].font = Font(bold=True, size=14, color=UNGU)
    ws["C1"] = ("Isi satu baris per temuan. Unggah tangkapan layarnya ke "
                "folder Drive, lalu tempel tautan berkasnya di kolom "
                "Link Capture.")
    ws["C1"].font = Font(size=9, italic=True, color="6E728C")

    # "Tombol" — sel bergaya tombol berisi HYPERLINK ke folder Drive.
    # Tombol sungguhan di Excel dijalankan makro, dan makro tidak bisa
    # mengunggah ke Drive tanpa kredensial API. Tautan justru lebih
    # tepat di sini: satu klik membuka foldernya di peramban, tempat
    # unggahannya memang dilakukan.
    tombol = ws["A2"]
    tombol.value = f'=HYPERLINK("{FOLDER_CAPTURE}", "BUKA FOLDER CAPTURE")'
    tombol.font = Font(bold=True, color=PUTIH, size=11)
    tombol.fill = PatternFill("solid", fgColor=UNGU)
    tombol.alignment = Alignment(horizontal="center", vertical="center")
    tombol.border = KOTAK
    ws.merge_cells("A2:C2")
    ws.row_dimensions[2].height = 24

    ws["D2"] = ("Unggah ke folder itu, beri nama sesuai ID temuan "
                "(DF-001.jpg), lalu Salin Tautan berkasnya dan tempel di "
                "kolom Link Capture.")
    ws["D2"].font = Font(size=9, color="6E728C")

    kepala = 4
    for i, (nama, lebar, _) in enumerate(KOLOM_DEFECT, start=1):
        sel = ws.cell(row=kepala, column=i, value=nama)
        sel.font = Font(bold=True, color=PUTIH, size=11)
        sel.fill = PatternFill("solid", fgColor=UNGU)
        sel.alignment = Alignment(vertical="center", horizontal="left",
                                  wrap_text=True)
        sel.border = KOTAK
        ws.column_dimensions[get_column_letter(i)].width = lebar
    ws.row_dimensions[kepala].height = 28

    kol_capture = get_column_letter(
        [k[0] for k in KOLOM_DEFECT].index("Link Capture") + 1)
    kol_buka = [k[0] for k in KOLOM_DEFECT].index("Buka") + 1

    awal, akhir = kepala + 1, kepala + BARIS_KOSONG
    for r in range(awal, akhir + 1):
        for c in range(1, len(KOLOM_DEFECT) + 1):
            sel = ws.cell(row=r, column=c)
            sel.border = KOTAK
            sel.alignment = Alignment(vertical="top", wrap_text=True)
            if r % 2 == 1:
                sel.fill = PatternFill("solid", fgColor=ABU)
        ws.cell(row=r, column=1).font = Font(bold=True)
        # Tautannya muncul sendiri begitu link-nya ditempel, dan tetap
        # kosong selama belum — sel berisi "Buka" yang tidak menuju ke
        # mana-mana lebih buruk daripada sel kosong.
        ws.cell(row=r, column=kol_buka).value = (
            f'=IF({kol_capture}{r}="","",'
            f'HYPERLINK({kol_capture}{r},"Buka"))'
        )
        ws.cell(row=r, column=kol_buka).font = Font(color=UNGU, underline="single")
        ws.cell(row=r, column=kol_buka).alignment = Alignment(
            horizontal="center", vertical="top")

    for i, (_, _, pilihan) in enumerate(KOLOM_DEFECT, start=1):
        if not pilihan:
            continue
        dv = DataValidation(type="list", formula1=f'"{pilihan}"',
                            allow_blank=True, showDropDown=False)
        ws.add_data_validation(dv)
        L = get_column_letter(i)
        dv.add(f"{L}{awal}:{L}{akhir}")

    ws.freeze_panes = f"C{kepala + 1}"
    ws.auto_filter.ref = (
        f"A{kepala}:{get_column_letter(len(KOLOM_DEFECT))}{akhir}")
    return ws, awal, akhir


def lembar_cara_capture(wb):
    """Satu halaman: cara mencatat temuan berikut tangkapan layarnya."""
    ws = wb.create_sheet("Cara Lampirkan Capture")
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 108

    baris = [
        ("Melampirkan tangkapan layar temuan", "judul"),
        ("", None),
        ("1.  Ambil tangkapan layar di HP saat temuannya terlihat.", None),
        ("2.  Klik tombol BUKA FOLDER CAPTURE di lembar Defect — folder "
         "Drive-nya terbuka di peramban.", None),
        ("3.  Unggah gambarnya ke sana, beri nama sesuai ID temuan: "
         "DF-001.jpg. Kalau butuh lebih dari satu, tambahkan keterangan: "
         "DF-001-sebelum.jpg, DF-001-sesudah.jpg.", None),
        ("4.  Klik kanan berkasnya di Drive - Bagikan - Salin tautan.", None),
        ("5.  Tempel tautannya di kolom Link Capture pada baris temuan itu. "
         "Kolom Buka di sebelahnya berubah sendiri jadi tautan yang bisa "
         "diklik.", None),
        ("", None),
        ("Pastikan aksesnya \"Siapa saja yang memiliki link\"", "judul"),
        ("", None),
        ("Tautan yang hanya bisa dibuka pemiliknya sama saja dengan tidak "
         "dilampirkan: yang membaca laporan ini nanti bukan orang yang "
         "mengunggahnya, dan dia akan berhenti di halaman minta izin.", None),
        ("", None),
        ("Kenapa gambarnya tidak ditempel langsung ke Excel", "judul"),
        ("", None),
        ("Satu tangkapan layar HP berukuran 300-800 KB. Tiga puluh temuan "
         "sudah cukup membuat berkas ini terlalu berat untuk dikirim lewat "
         "chat mana pun — dan berkas yang tidak bisa dikirim tidak dipakai "
         "siapa-siapa.", None),
        ("", None),
        ("Kenapa tombolnya berupa tautan, bukan makro", "judul"),
        ("", None),
        ("Tombol yang menjalankan perintah di Excel butuh makro, dan makro "
         "tidak bisa mengunggah ke Google Drive tanpa kredensial API yang "
         "harus dititipkan di dalam berkasnya — kredensial di dalam berkas "
         "yang dikirim ke banyak orang sama saja dengan diumumkan.", None),
        ("", None),
        ("Selain itu makro hanya hidup di berkas .xlsm: mati di Google "
         "Sheets, mati di Numbers, dan ditolak sebagian penyetelan keamanan "
         "kantor. Tautan bekerja di semuanya, dan satu klik membuka folder "
         "tempat unggahannya memang dilakukan.", None),
    ]

    r = 1
    for teks, jenis in baris:
        sel = ws.cell(row=r, column=2, value=teks)
        if jenis == "judul":
            sel.font = Font(bold=True, size=12, color=UNGU)
        else:
            sel.font = Font(size=10.5)
            sel.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
    return ws


def main():
    kasus = baca(SUMBER) + e2e(SUMBER)

    wb = Workbook()
    wb.remove(wb.active)

    ws_kasus = lembar_kasus(wb, kasus)
    _, awal_df, akhir_df = (None, 0, 0)
    ws_defect, awal_df, akhir_df = lembar_defect(wb)
    lembar_prasyarat(wb, SUMBER)
    lembar_cara_capture(wb)
    ws_rekap, baris_temuan = lembar_rekap(wb, ws_kasus, kasus)
    rekap_temuan(ws_rekap, baris_temuan, awal_df, akhir_df)

    wb.move_sheet("Rekap", offset=-wb.sheetnames.index("Rekap"))
    wb.active = 0
    wb.save(TUJUAN)
    print(f"ditulis: {TUJUAN}")
    print(f"  {len(kasus)} kasus, {len(set(k[0] for k in kasus))} modul")
    print(f"  lembar Defect siap {BARIS_KOSONG} baris")


if __name__ == "__main__":
    main()
